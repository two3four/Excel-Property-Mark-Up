
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font
import io
import re
import math
import requests

st.set_page_config(page_title="Excel Property Mark-Up Tool", layout="centered")

def base_rate_index(units_per_building):
    if units_per_building <= 1.99:
        return 1
    elif units_per_building <= 3.99:
        return 2
    elif units_per_building <= 9.99:
        return 3
    elif units_per_building <= 19.99:
        return 4
    else:
        return 5

management_companies = [
    "AMC", "American Landmark", "Asset living", "AvalonBay", "Avenue5", "Beacon", "Bell", "Berkshire", "BH Management", "Bozzuto",
    "Bridge", "Bridge Investment Group", "Brookside", "Bryten", "Camden", "Cirrus", "CMC", "CONAM", "Continental", "Cornerstone",
    "Cortland", "Cushman & Wakefield", "Dominium", "Drucker & Falk", "Edward Rose & Sons", "Elmington", "Elon", "Envolve", "Equity Residential",
    "Essex", "Fairfield", "First Communities Management", "FPI", "Franklin Group", "Gateway", "Georgia MLS Real Estate", "Greystar",
    "Hawthorne Residential Partners Llc", "Highmark", "Irvine", "John Stewart", "Landmark", "MAA", "Mark", "Mercy Housing", "Millennia",
    "Mission Rock", "Monarch", "Morgan Properties", "Moss & Company", "Pedcor Companies", "Pegasus", "Pinnacle", "PPM", "Premier",
    "Princeton Management", "RAM", "Rangewater", "Redwood", "Related Management", "Royal American Companies", "RPM", "S.L. Nusbaum", "SPM",
    "Tm Associates", "TMO", "Trinity", "UAG", "UDR", "UMH Properties", "Village Green", "Weidner", "Westdale", "Westminster", "Wilhoit",
    "Willow Bridge", "Windsor", "WINN", "Woda", "Yes! Communities", "ZRS"
]

def app():
    st.title("📊 Excel Property Mark-Up Tool")

    url = "https://github.com/two3four/Excel-Property-Mark-Up/raw/main/Calculator.xlsx"
    response = requests.get(url)
    uploaded_file = io.BytesIO(response.content)

    if uploaded_file:
        wb = load_workbook(uploaded_file)
        ws = wb.active

        ws["C1"] = datetime.today().strftime("%A, %B %d, %Y")

        st.subheader("🔧 Basic Info")
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Your Name", value="", key="name", autocomplete="off")
        with col2:
            prop = st.text_input("Property Name & Address", value="", key="property", autocomplete="off")

        display_text = f"{prop} ({name})" if prop and name else prop or name or ""
        output_filename = "Updated_Calculator.xlsx"
        if display_text.strip():
            ws["C3"] = display_text
        if prop.strip():
            safe_filename = re.sub(r'[\\/:*?"<>|]', '', prop).strip()
            output_filename = f"{safe_filename}.xlsx"

        st.subheader("🏢 Building Details")
        col3, col4 = st.columns(2)
        with col3:
            buildings = st.number_input("Number of Buildings", min_value=1, value=1)
            ws["B13"] = buildings
        with col4:
            units = st.number_input("Number of Units", min_value=1, value=1)
            ws["B14"] = units

        if buildings > 0:
            ws["C14"] = base_rate_index(units / buildings)

        st.subheader("🌳 Tree & Structure Info")
        ws["E14"] = {"No": 1, "Light": 2, "Moderate": 3, "Heavy": 4}[st.selectbox("Tree Coverage", ["No", "Light", "Moderate", "Heavy"])]
        ws["G14"] = {"Broadleaf": 1, "Conifer": 2, "Mixed": 3}[st.selectbox("Tree Type", ["Broadleaf", "Conifer", "Mixed"])]

        story = st.selectbox("Number of Stories", [
            "1", "2", "3", "4+ with no roof access (lift/ScyVac)", "High-rise (anything 3+ with roof access)"
        ])
        ws["I14"] = {"1": 1, "2": 2, "3": 3,
                     "4+ with no roof access (lift/ScyVac)": 4,
                     "High-rise (anything 3+ with roof access)": 5}[story]

        ws["K14"] = {"Regular": 1, "Irregular": 2, "Complex": 3, "Very Complex": 4}[st.selectbox("Complexity", [
            "Regular", "Irregular", "Complex", "Very Complex"
        ])]
        ws["M14"] = {"Walkable": 1, "Partially": 2, "Unwalkable": 3}[st.selectbox("Walkability", ["Walkable", "Partially", "Unwalkable"])]
        ws["O14"] = {"No Obstacles": 1, "Some obstacles": 2, "Many obstacles": 3, "Very Complex": 4}[
            st.selectbox("Balconies/patios for unwalkable", ["No Obstacles", "Some obstacles", "Many obstacles", "Very Complex"])
        ]

        st.subheader("🚗 Parking Info")
        g_trees = st.number_input("Garages with Trees", min_value=0, value=0)
        g_none = st.number_input("Garages without Trees", min_value=0, value=0)
        c_trees = st.number_input("Carports with Trees", min_value=0, value=0)
        c_none = st.number_input("Carports without Trees", min_value=0, value=0)

        ws["H19"] = g_trees
        ws["H21"] = g_none
        ws["H26"] = c_trees
        ws["H28"] = c_none

        ws["B19"] = math.ceil(g_trees + g_none / 4)
        ws["E19"] = math.ceil(c_trees + c_none / 4)

        st.subheader("🏢 Management Company Discount")
        selected_company = st.selectbox("Select a Company", ["None"] + management_companies)
        if selected_company != "None":
            ws["B16"] = 1

        # BULK DISCOUNT LOGIC - R19 (merged R19:S20) → F27:F54
        r19_val = ws["R19"].value
        if isinstance(r19_val, (int, float)):
            for row in range(27, 55):  # D27 to D54 (actual data rows)
                cell_val = ws[f"D{row}"].value
                if cell_val:
                    text = str(cell_val).replace(",", "").strip()
                    if "to" in text:
                        parts = text.split("to")
                        low, high = int(parts[0]), int(parts[1])
                    elif "and up" in text or "+" in text:
                        low, high = int(text.split()[0]), float("inf")
                    else:
                        continue
                    if low <= r19_val <= high:
                        ws[f"F{row}"] = 1
                        break

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button("📥 Download Updated Excel File", output.getvalue(), file_name=output_filename)

if __name__ == "__main__":
    app()
